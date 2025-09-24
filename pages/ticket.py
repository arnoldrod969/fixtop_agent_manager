import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from database import db_manager
from permissions import PermissionManager
import plotly.express as px
import plotly.graph_objects as go
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

if "logged_in" not in st.session_state or not st.session_state.logged_in:
    st.switch_page("app.py")  # Redirect to home/login if not connected

# Check page access permissions
if not PermissionManager.check_page_access("ticket_page"):
    PermissionManager.show_access_denied("You do not have access to ticket management.")

# Page configuration
st.set_page_config(page_title="Ticket Management", page_icon="🎫", layout="wide")

# Custom CSS
st.markdown(
    """
<style>
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin: 0.5rem 0;
    }
    .ticket-card {
        border: 1px solid #e0e0e0;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
        background: white;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .status-active {
        color: #28a745;
        font-weight: bold;
    }
    .status-inactive {
        color: #dc3545;
        font-weight: bold;
    }
</style>
""",
    unsafe_allow_html=True,
)

# Main title
st.title("🎫 Ticket Management")
st.markdown("---")


# Utility functions
@st.cache_data(ttl=60)
def load_tickets():
    """Loads all tickets from the database"""
    return db_manager.get_all_problems()


@st.cache_data(ttl=60)
def load_ticket_stats():
    """Loads ticket statistics"""
    return db_manager.get_problem_stats()


@st.cache_data(ttl=60)
def load_domains():
    """Loads all domains from the database"""
    try:
        with db_manager.get_connection() as conn:
            cursor = conn.execute(
                """
                SELECT id, name FROM craft 
                WHERE is_active = 1 
                ORDER BY name
            """
            )
            return [dict(row) for row in cursor.fetchall()]
    except Exception as e:
        st.error(f"Error loading domains: {str(e)}")
        return []


@st.cache_data(ttl=60)
def load_specialties_by_domain(domain_id):
    """Loads specialties for the selected domain"""
    if not domain_id:
        return []
    try:
        with db_manager.get_connection() as conn:
            cursor = conn.execute(
                """
                SELECT id, name FROM speciality 
                WHERE craft_id = ? AND is_active = 1 
                ORDER BY name
            """,
                (domain_id,),
            )
            return [dict(row) for row in cursor.fetchall()]
    except Exception as e:
        st.error(f"Error loading specialties: {str(e)}")
        return []


@st.cache_data(ttl=60)
def load_agents():
    """Loads all active agents from the database"""
    try:
        users = db_manager.get_all_users()
        if users:
            # Filter only active agents
            agents = [user for user in users if user['role_name'] == 'agent' and user['is_active'] == 1]
            return agents
        return []
    except Exception as e:
        st.error(f"Error loading agents: {str(e)}")
        return []


def clear_cache():
    """Clears cache to refresh data"""
    st.cache_data.clear()


def export_to_csv(dataframe):
    """Export dataframe to CSV format"""
    try:
        return dataframe.to_csv(index=False).encode("utf-8")
    except Exception as e:
        st.error(f"Error exporting to CSV: {str(e)}")
        return None


def export_to_pdf(dataframe, title="Ticket Statistics Report"):
    """Export dataframe to PDF format"""
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=50,
            bottomMargin=30,
        )

        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=14,
            spaceAfter=20,
            alignment=1,
        )

        title_para = Paragraph(title, title_style)
        elements.append(title_para)
        elements.append(Spacer(1, 8))

        date_para = Paragraph(
            f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles["Normal"],
        )
        elements.append(date_para)
        elements.append(Spacer(1, 12))

        data = [list(dataframe.columns)]
        for _, row in dataframe.iterrows():
            processed_row = []
            for i, cell in enumerate(row):
                cell_str = str(cell)
                if i == 0:  # ID
                    processed_row.append(cell_str[:8])
                elif i == 1:  # Customer Name
                    processed_row.append(cell_str[:15])
                elif i == 2:  # Phone
                    processed_row.append(cell_str[:12])
                elif i == 3:  # Problem Description
                    processed_row.append(
                        cell_str[:25] + "..." if len(cell_str) > 25 else cell_str
                    )
                else:
                    processed_row.append(cell_str[:15])
            data.append(processed_row)

        available_width = A4[0] - 60
        col_count = len(dataframe.columns)
        col_width = available_width / col_count

        table = Table(data, colWidths=[col_width] * col_count)

        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("TOPPADDING", (0, 0), (-1, 0), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )

        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    except Exception as e:
        st.error(f"Error exporting to PDF: {str(e)}")
        return None


def export_to_excel(dataframe, title="Ticket Statistics Report"):
    """Export dataframe to Excel format"""
    try:
        buffer = io.BytesIO()

        wb = Workbook()
        ws = wb.active
        ws.title = "Ticket Statistics"

        ws["A1"] = title
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        max_col = len(dataframe.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

        ws["A2"] = f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(size=10, italic=True)

        start_row = 4

        for col_idx, column_name in enumerate(dataframe.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        for row_idx, (_, row) in enumerate(dataframe.iterrows(), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value))

        for col_idx in range(1, len(dataframe.columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0

            for row_idx in range(start_row, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_length = len(str(cell_value))
                    if cell_length > max_length:
                        max_length = cell_length

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    except Exception as e:
        st.error(f"Error exporting to Excel: {str(e)}")
        return None


# Create tabs based on permissions
available_tabs = PermissionManager.get_available_tabs("ticket_page")

# Dynamic tab creation according to permissions
tabs = st.tabs(available_tabs)

# Tab variable initialization
tab1 = tab2 = tab3 = tab4 = tab5 = None

# Tab assignment according to their content
for i, tab_name in enumerate(available_tabs):
    if "List" in tab_name:
        tab1 = tabs[i]
    elif "Add" in tab_name:
        tab2 = tabs[i]
    elif "Edit" in tab_name:
        tab3 = tabs[i]
    elif "Delete" in tab_name:
        tab4 = tabs[i]
    elif "Statistics" in tab_name:
        tab5 = tabs[i]

# ==================== LIST TAB ====================
with tab1:
    st.header("Ticket List")

    # Filters
    col1, col2, col3 = st.columns(3)

    with col1:
        search_customer = st.text_input(
            "🔍 Search by customer name", key="search_customer"
        )

    with col2:
        search_phone = st.text_input("📱 Search by phone", key="search_phone")

    with col3:
        if st.button("🔄 Refresh", key="refresh_list"):
            clear_cache()
            st.rerun()

    # Data loading
    tickets = load_tickets()

    if tickets:
        # Data filtering
        filtered_tickets = tickets

        if search_customer:
            filtered_tickets = [
                t
                for t in filtered_tickets
                if search_customer.lower() in t["customer_name"].lower()
            ]

        if search_phone:
            filtered_tickets = [
                t for t in filtered_tickets if search_phone in t["customer_phone"]
            ]

        # Results display
        st.info(f"📊 {len(filtered_tickets)} ticket(s) found")

        # Ticket table
        if filtered_tickets:
            df = pd.DataFrame(filtered_tickets)

            # Select columns to display
            display_columns = [
                "id",
                "customer_name",
                "customer_phone",
                "problem_desc",
                "created_by_name",
                "created_at",
            ]

            # Column renaming for display
            column_names = {
                "id": "ID",
                "customer_name": "Customer Name",
                "customer_phone": "Phone",
                "problem_desc": "Problem Description",
                "created_by_name": "Created by",
                "created_at": "Creation Date",
            }

            df_display = df[display_columns].rename(columns=column_names)

            # Dataframe display configuration
            st.dataframe(
                df_display,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Problem Description": st.column_config.TextColumn(width="large"),
                    "Creation Date": st.column_config.DatetimeColumn(
                        format="DD/MM/YYYY HH:mm"
                    ),
                },
            )
        else:
            st.warning("No tickets match the search criteria.")
    else:
        st.info("No tickets found in the database.")

# ==================== ADD TAB ====================
if tab2 is not None:  # Only if user has permissions
    with tab2:
        st.header("Add a New Ticket")

        # Load domains for the form
        domains = load_domains()

        # Initialize session variables for the form
        if "form_customer_name" not in st.session_state:
            st.session_state.form_customer_name = ""
        if "form_customer_phone" not in st.session_state:
            st.session_state.form_customer_phone = ""
        if "form_problem_desc" not in st.session_state:
            st.session_state.form_problem_desc = ""
        if "form_payment" not in st.session_state:
            st.session_state.form_payment = "No"
        if "form_amount" not in st.session_state:
            st.session_state.form_amount = 0
        if "form_domain" not in st.session_state:
            st.session_state.form_domain = "Select a domain..."
        if "form_specialties" not in st.session_state:
            st.session_state.form_specialties = []
        if "previous_domain_id" not in st.session_state:
            st.session_state.previous_domain_id = None

        col1, col2 = st.columns(2)

        with col1:
            customer_name = st.text_input(
                "Customer Name *",
                value=st.session_state.form_customer_name,
                key="add_customer_name",
            )
            customer_phone = st.text_input(
                "Customer Phone *",
                value=st.session_state.form_customer_phone,
                key="add_customer_phone",
            )

            # Payment field
            payment = st.selectbox(
                "Payment *",
                options=["No", "Yes"],
                index=0 if st.session_state.form_payment == "No" else 1,
                key="add_payment",
                help="Has the customer made a payment?",
            )

            # Immediate amount update when payment = "No"
            if payment == "No":
                st.session_state.form_amount = 0

            # Amount field (conditional) - displays immediately
            amount = None
            if payment == "Yes":
                amount = st.number_input(
                    "Amount (₦) *",
                    min_value=0,
                    step=1,
                    value=int(st.session_state.form_amount),
                    key="add_amount",
                    help="Payment amount in naira",
                )
            else:
                # Display amount at 0 when payment = "No"
                st.number_input(
                    "Amount (₦)",
                    value=0,
                    disabled=True,
                    help="Amount is 0 because no payment was made",
                )

        with col2:
            # Domain field (simple selection)
            if domains:
                domain_options = {d["name"]: d["id"] for d in domains}
                domain_list = ["Select a domain..."] + list(domain_options.keys())

                try:
                    domain_index = domain_list.index(st.session_state.form_domain)
                except ValueError:
                    domain_index = 0

                selected_domain = st.selectbox(
                    "Domain *",
                    options=domain_list,
                    index=domain_index,
                    key="add_domain",
                    help="Select the domain related to the problem",
                )
                selected_domain_id = (
                    domain_options.get(selected_domain)
                    if selected_domain != "Select a domain..."
                    else None
                )

                # Detect if domain changed to reset specialties
                if "previous_domain_id" not in st.session_state:
                    st.session_state.previous_domain_id = None

                if st.session_state.previous_domain_id != selected_domain_id:
                    st.session_state.form_specialties = []
                    st.session_state.previous_domain_id = selected_domain_id

            else:
                st.warning("No domains available")
                selected_domain = None
                selected_domain_id = None

            # Specialties field (multi-selection dependent on selected domain)
            selected_specialties = []
            selected_specialty_ids = []
            if selected_domain_id:
                specialties = load_specialties_by_domain(selected_domain_id)
                if specialties:
                    specialty_options = {s["name"]: s["id"] for s in specialties}

                    # Filter default specialties to keep only those that exist for this domain
                    valid_default_specialties = [
                        spec
                        for spec in st.session_state.form_specialties
                        if spec in specialty_options.keys()
                    ]

                    selected_specialties = st.multiselect(
                        "Specialties",
                        options=list(specialty_options.keys()),
                        default=valid_default_specialties,
                        key="add_specialties",
                        help="Select the relevant specialties",
                    )
                    selected_specialty_ids = [
                        specialty_options[name] for name in selected_specialties
                    ]

                    # Immediate session update to avoid double-click
                    st.session_state.form_specialties = selected_specialties
                else:
                    st.info("No specialties available for this domain")
            else:
                st.info("First select a domain to see specialties")

            # Immediate update of selected domain
            if selected_domain:
                st.session_state.form_domain = selected_domain

        problem_desc = st.text_area(
            "Problem Description *",
            height=150,
            value=st.session_state.form_problem_desc,
            key="add_problem_desc",
            help="Describe in detail the problem encountered by the customer",
        )

        # Action buttons
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])

        with col_btn1:
            if st.button("➕ Create Ticket", type="primary", key="create_ticket_btn"):
                # Validate required fields
                errors = []
                if not customer_name:
                    errors.append("Customer name")
                if not customer_phone:
                    errors.append("Customer phone")
                if not problem_desc:
                    errors.append("Problem description")
                if not selected_domain_id:
                    errors.append("A domain")
                if payment == "Yes" and (amount is None or amount <= 0):
                    errors.append("Payment amount (must be greater than 0)")

                if errors:
                    st.error(
                        f"⚠️ The following fields are required: {', '.join(errors)}"
                    )
                else:
                    # Create ticket with new fields
                    try:
                        with db_manager.get_connection() as conn:
                            # Insert main ticket
                            cursor = conn.execute(
                                """
                                 INSERT INTO problems (customer_name, customer_phone, problem_desc, 
                                                     is_paid, amount, craft_ids, speciality_ids, created_by , updated_by)
                                 VALUES (?, ?, ?, ?, ?, ?, ?, ? , ?)
                             """,
                                (
                                    customer_name.strip(),
                                    customer_phone.strip(),
                                    problem_desc.strip(),
                                    1 if payment == "Yes" else 0,
                                    amount if payment == "Yes" else 0,
                                    (
                                        str(selected_domain_id)
                                        if selected_domain_id
                                        else None
                                    ),
                                    (
                                        ",".join(map(str, selected_specialty_ids))
                                        if selected_specialty_ids
                                        else None
                                    ),
                                    st.session_state.user_id,
                                    st.session_state.user_id,
                                ),
                            )

                            problem_id = cursor.lastrowid

                            conn.commit()
                            st.success(
                                f"✅ Ticket created successfully (ID: {problem_id})"
                            )

                            # Reset form
                            st.session_state.form_customer_name = ""
                            st.session_state.form_customer_phone = ""
                            st.session_state.form_problem_desc = ""
                            st.session_state.form_payment = "No"
                            st.session_state.form_amount = 0
                            st.session_state.form_domain = "Select a domain..."
                            st.session_state.form_specialties = []
                            st.session_state.previous_domain_id = None

                            clear_cache()
                            st.rerun()

                    except Exception as e:
                        st.error(f"❌ Error during creation: {str(e)}")

        with col_btn2:
            if st.button("🔄 Reset", key="reset_form_btn"):
                # Reset all form fields
                st.session_state.form_customer_name = ""
                st.session_state.form_customer_phone = ""
                st.session_state.form_problem_desc = ""
                st.session_state.form_payment = "No"
                st.session_state.form_amount = 0
                st.session_state.form_domain = "Select a domain..."
                st.session_state.form_specialties = []
                st.session_state.previous_domain_id = None
                st.rerun()

        # Update session variables (only for fields not managed immediately)
        st.session_state.form_customer_name = customer_name
        st.session_state.form_customer_phone = customer_phone
        st.session_state.form_problem_desc = problem_desc
        st.session_state.form_payment = payment
        if amount is not None:
            st.session_state.form_amount = amount

# ==================== EDIT TAB ====================
if tab3 is not None:  # Only if user has permissions
    with tab3:
        st.header("Edit a Ticket")

        tickets = load_tickets()

        if tickets:
            # Select ticket to edit
            ticket_options = {
                f"#{t['id']} - {t['customer_name']} ({t['customer_phone']})": t["id"]
                for t in tickets
            }

            selected_ticket_key = st.selectbox(
                "Select a ticket to edit",
                options=list(ticket_options.keys()),
                key="modify_ticket_select",
            )

            if selected_ticket_key:
                ticket_id = ticket_options[selected_ticket_key]
                ticket = db_manager.get_problem_by_id(ticket_id)

                if ticket:
                    with st.form("modify_ticket_form"):
                        col1, col2 = st.columns(2)

                        with col1:
                            new_customer_name = st.text_input(
                                "Customer Name",
                                value=ticket["customer_name"],
                                key="modify_customer_name",
                            )
                            new_customer_phone = st.text_input(
                                "Customer Phone",
                                value=ticket["customer_phone"],
                                key="modify_customer_phone",
                            )

                        with col2:
                            st.info(
                                f"**Created by:** {ticket['created_by_name'] or 'Unknown'}"
                            )
                            st.info(f"**Creation date:** {ticket['created_at']}")

                        new_problem_desc = st.text_area(
                            "Problem Description",
                            value=ticket["problem_desc"],
                            height=150,
                            key="modify_problem_desc",
                        )

                        submitted = st.form_submit_button("💾 Update", type="primary")

                        if submitted:
                            if (
                                not new_customer_name
                                or not new_customer_phone
                                or not new_problem_desc
                            ):
                                st.error("⚠️ All fields are required.")
                            else:
                                # Update ticket
                                success, message = db_manager.update_problem(
                                    problem_id=ticket_id,
                                    customer_name=new_customer_name.strip(),
                                    customer_phone=new_customer_phone.strip(),
                                    problem_desc=new_problem_desc.strip(),
                                    updated_by=st.session_state.user_id,
                                )

                                if success:
                                    st.success(f"✅ {message}")
                                    clear_cache()
                                    st.rerun()
                                else:
                                    st.error(f"❌ {message}")
        else:
            st.info("No tickets available for editing.")

# ==================== DELETE TAB ====================
if tab4 is not None:  # Only if user has permissions
    with tab4:
        st.header("Delete a Ticket")
        st.warning(
            "⚠️ Warning: This action will mark the ticket as inactive (logical deletion)."
        )

        tickets = load_tickets()

        if tickets:
            # Ticket selection for deletion
            ticket_options = {
                f"#{t['id']} - {t['customer_name']} ({t['customer_phone']})": t["id"]
                for t in tickets
            }

            selected_ticket_key = st.selectbox(
                "Select a ticket to delete",
                options=list(ticket_options.keys()),
                key="delete_ticket_select",
            )

            if selected_ticket_key:
                ticket_id = ticket_options[selected_ticket_key]
                ticket = db_manager.get_problem_by_id(ticket_id)

                if ticket:
                    # Display ticket details
                    st.markdown("### Details of the ticket to delete:")

                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**ID:** {ticket['id']}")
                        st.write(f"**Customer:** {ticket['customer_name']}")
                        st.write(f"**Phone:** {ticket['customer_phone']}")

                    with col2:
                        st.write(
                            f"**Created by:** {ticket['created_by_name'] or 'Unknown'}"
                        )
                        st.write(f"**Date:** {ticket['created_at']}")

                    st.write(f"**Problem:** {ticket['problem_desc']}")

                    # Deletion confirmation
                    col1, col2, col3 = st.columns([1, 1, 2])

                    with col1:
                        if st.button(
                            "🗑️ Confirm Deletion", type="primary", key="confirm_delete"
                        ):
                            success, message = db_manager.delete_problem(ticket_id)

                            if success:
                                st.success(f"✅ {message}")
                                clear_cache()
                                st.rerun()
                            else:
                                st.error(f"❌ {message}")

                    with col2:
                        if st.button("❌ Cancel", key="cancel_delete"):
                            st.rerun()
        else:
            st.info("No tickets available for deletion.")

# ==================== STATISTICS TAB ====================
if tab5 is not None:
    with tab5:
        st.header("Ticket Statistics")

        # Load tickets data
        tickets = load_tickets()

        if tickets:
            # Convert to DataFrame for easier manipulation
            tickets_df = pd.DataFrame(tickets)

            # Advanced Filters Section
            st.subheader("🔍 Advanced Filters")

            # Create filter columns - now 6 columns to include "created by" filter
            filter_col1, filter_col2, filter_col3, filter_col4, filter_col5, filter_col6 = (
                st.columns(6)
            )

            with filter_col1:
                search_ticket = st.text_input(
                    "🔍 Search",
                    placeholder="Customer name, phone, description...",
                    key="stats_search_ticket",
                )

            with filter_col2:
                # Payment status filter
                payment_filter = st.selectbox(
                    "💰 Payment Status",
                    ["All", "Paid", "Unpaid"],
                    help="Filter by payment status",
                    key="stats_payment_filter",
                )

            with filter_col3:
                # Domain filter - Updated for multiselect
                domains = load_domains()
                if domains:
                    domain_names = [d["name"] for d in domains]
                    domain_filter = st.multiselect(
                        "🏢 Domain",
                        options=domain_names,
                        help="Select one or more domains",
                        key="stats_domain_filter",
                    )
                else:
                    domain_filter = []

            with filter_col4:
                # Specialty filter - Updated for multiselect and multiple domains
                specialty_options = []
                if domain_filter:  # Only if domains are selected
                    selected_domain_ids = [
                        d["id"] for d in domains if d["name"] in domain_filter
                    ]
                    all_specialties = []
                    for domain_id in selected_domain_ids:
                        specialties = load_specialties_by_domain(domain_id)
                        all_specialties.extend(specialties)

                    # Remove duplicates while preserving order
                    seen = set()
                    unique_specialties = []
                    for specialty in all_specialties:
                        if specialty["name"] not in seen:
                            unique_specialties.append(specialty)
                            seen.add(specialty["name"])

                    specialty_options = [s["name"] for s in unique_specialties]

                specialty_filter = st.multiselect(
                    "🎯 Specialty",
                    options=specialty_options,
                    help="Select one or more specialties (depends on selected domains)",
                    key="stats_specialty_filter",
                )

            with filter_col5:
                # Created by filter (multiselect)
                agents = load_agents()
                if agents:
                    agent_names = [agent['name'] for agent in agents]
                    created_by_filter = st.multiselect(
                        "👤 Created by",
                        options=agent_names,
                        help="Select one or more agents",
                        key="stats_created_by_filter",
                    )
                else:
                    created_by_filter = []

            with filter_col6:
                # Date filter
                date_filter = st.selectbox(
                    "📅 Date Filter",
                    ["All", "Creation Date", "Modification Date"],
                    help="Filter by date",
                    key="stats_date_filter",
                )

            # Date range inputs
            start_date = None
            end_date = None
            if date_filter != "All":
                date_col1, date_col2 = st.columns(2)
                
                # Calculer des valeurs par défaut intelligentes
                if not tickets_df.empty:
                    min_date = tickets_df['created_at'].min() if date_filter == "Creation Date" else tickets_df.get('updated_at', tickets_df['created_at']).min()
                    max_date = tickets_df['created_at'].max() if date_filter == "Creation Date" else tickets_df.get('updated_at', tickets_df['created_at']).max()
                    
                    # Convert to date if datetime
                    try:
                        min_date = pd.to_datetime(min_date).date()
                        max_date = pd.to_datetime(max_date).date()
                    except:
                        from datetime import date, timedelta
                        min_date = date.today() - timedelta(days=30)
                        max_date = date.today()
                else:
                    from datetime import date, timedelta
                    min_date = date.today() - timedelta(days=30)
                    max_date = date.today()
                
                with date_col1:
                    start_date = st.date_input("From", value=min_date, key="statistics_start_date")
                with date_col2:
                    end_date = st.date_input("To", value=max_date, key="statistics_end_date")

            # Apply filters to tickets data
            filtered_tickets = tickets_df.copy()

            # Search filter
            if search_ticket:
                search_mask = (
                    filtered_tickets["customer_name"].str.contains(
                        search_ticket, case=False, na=False
                    )
                    | filtered_tickets["customer_phone"].str.contains(
                        search_ticket, case=False, na=False
                    )
                    | filtered_tickets["problem_desc"].str.contains(
                        search_ticket, case=False, na=False
                    )
                )
                filtered_tickets = filtered_tickets[search_mask]

            # Payment filter
            if payment_filter != "All":
                if payment_filter == "Paid":
                    filtered_tickets = filtered_tickets[
                        filtered_tickets["is_paid"] == 1
                    ]
                elif payment_filter == "Unpaid":
                    filtered_tickets = filtered_tickets[
                        filtered_tickets["is_paid"] == 0
                    ]

            # Domain filter (multiselect)
            if domain_filter:
                selected_domain_ids = [
                    d["id"] for d in domains if d["name"] in domain_filter
                ]
                if selected_domain_ids:
                    # Filter tickets that match any of the selected domains
                    domain_mask = (
                        filtered_tickets["craft_ids"]
                        .astype(str)
                        .apply(
                            lambda x: (
                                any(
                                    str(domain_id) == str(x)
                                    for domain_id in selected_domain_ids
                                )
                                if pd.notna(x)
                                else False
                            )
                        )
                    )
                    filtered_tickets = filtered_tickets[domain_mask]

            # Specialty filter (multiselect and dependent on domains)
            if specialty_filter and domain_filter:
                # Get specialty IDs from selected specialty names
                selected_specialty_ids = []
                selected_domain_ids = [
                    d["id"] for d in domains if d["name"] in domain_filter
                ]
                for domain_id in selected_domain_ids:
                    specialties = load_specialties_by_domain(domain_id)
                    for specialty_name in specialty_filter:
                        specialty = next(
                            (s for s in specialties if s["name"] == specialty_name),
                            None,
                        )
                        if specialty and specialty["id"] not in selected_specialty_ids:
                            selected_specialty_ids.append(specialty["id"])

                if selected_specialty_ids:
                    # Filter tickets that match any of the selected specialties
                    specialty_mask = (
                        filtered_tickets["speciality_ids"]
                        .astype(str)
                        .apply(
                            lambda x: (
                                any(
                                    str(specialty_id) in str(x)
                                    for specialty_id in selected_specialty_ids
                                )
                                if pd.notna(x)
                                else False
                            )
                        )
                    )
                    filtered_tickets = filtered_tickets[specialty_mask]

            # Created by filter (multiselect)
            if created_by_filter:
                # Get agent IDs from selected agent names
                selected_agent_ids = []
                for agent_name in created_by_filter:
                    agent = next(
                        (a for a in agents if a['name'] == agent_name),
                        None,
                    )
                    if agent:
                        selected_agent_ids.append(agent["id"])

                if selected_agent_ids:
                    # Filter tickets that match any of the selected agents
                    created_by_mask = filtered_tickets["created_by"].isin(selected_agent_ids)
                    filtered_tickets = filtered_tickets[created_by_mask]

            # Date filter
            if date_filter == "Creation Date" and start_date is not None and end_date is not None:
                filtered_tickets["created_at_dt"] = pd.to_datetime(filtered_tickets["created_at"])
                filtered_tickets = filtered_tickets[
                    (filtered_tickets["created_at_dt"].dt.date >= start_date) &
                    (filtered_tickets["created_at_dt"].dt.date <= end_date)
                ]
                filtered_tickets = filtered_tickets.drop("created_at_dt", axis=1)
            elif date_filter == "Modification Date" and start_date is not None and end_date is not None:
                # Check if updated_at column exists, if not use created_at
                if 'updated_at' in filtered_tickets.columns:
                    filtered_tickets["updated_at_dt"] = pd.to_datetime(filtered_tickets["updated_at"])
                    filtered_tickets = filtered_tickets[
                        (filtered_tickets["updated_at_dt"].dt.date >= start_date) &
                        (filtered_tickets["updated_at_dt"].dt.date <= end_date)
                    ]
                    filtered_tickets = filtered_tickets.drop("updated_at_dt", axis=1)
                else:
                    # Fallback to created_at if updated_at doesn't exist
                    filtered_tickets["created_at_dt"] = pd.to_datetime(filtered_tickets["created_at"])
                    filtered_tickets = filtered_tickets[
                        (filtered_tickets["created_at_dt"].dt.date >= start_date) &
                        (filtered_tickets["created_at_dt"].dt.date <= end_date)
                    ]
                    filtered_tickets = filtered_tickets.drop("created_at_dt", axis=1)

            # Key Metrics
            st.subheader("📈 Key Metrics")

            # Calculate metrics
            total_tickets = len(filtered_tickets)
            paid_tickets = (
                len(filtered_tickets[filtered_tickets["is_paid"] == 1])
                if not filtered_tickets.empty
                else 0
            )
            total_amount = (
                filtered_tickets["amount"].sum() if not filtered_tickets.empty else 0
            )
            avg_amount = (
                filtered_tickets[filtered_tickets["amount"] > 0]["amount"].mean()
                if not filtered_tickets.empty
                else 0
            )

            col1, col2, col3, col4 = st.columns(4)

            with col1:
                st.metric("📊 Total Tickets", total_tickets)
            with col2:
                st.metric("💰 Paid Tickets", paid_tickets)
            with col3:
                st.metric("💵 Total Amount", f"₦{total_amount:,.0f}")
            with col4:
                st.metric(
                    "📈 Avg Amount", f"₦{avg_amount:.0f}" if avg_amount > 0 else "₦0"
                )

            # Enhanced data table
            st.subheader("📋 Detailed Ticket List")

            if not filtered_tickets.empty:
                # Prepare enhanced data
                enhanced_data = []
                for _, ticket in filtered_tickets.iterrows():
                    # Get domain name
                    domain_name = "N/A"
                    if ticket.get("craft_ids"):
                        try:
                            domain_id = int(ticket["craft_ids"])
                            domain = next(
                                (d for d in domains if d["id"] == domain_id), None
                            )
                            if domain:
                                domain_name = domain["name"]
                        except:
                            pass

                    # Get specialty names
                    specialty_names = []
                    if ticket.get("speciality_ids"):
                        try:
                            specialty_ids = ticket["speciality_ids"].split(",")
                            for spec_id in specialty_ids:
                                if spec_id.strip():
                                    # Load specialties for the domain to get names
                                    if ticket.get("craft_ids"):
                                        domain_id = int(ticket["craft_ids"])
                                        specialties = load_specialties_by_domain(
                                            domain_id
                                        )
                                        specialty = next(
                                            (
                                                s
                                                for s in specialties
                                                if s["id"] == int(spec_id.strip())
                                            ),
                                            None,
                                        )
                                        if specialty:
                                            specialty_names.append(specialty["name"])
                        except:
                            pass

                    # Format creation date
                    try:
                        created_date = pd.to_datetime(ticket["created_at"]).strftime(
                            "%d/%m/%Y %H:%M"
                        )
                    except:
                        created_date = ticket["created_at"]

                    # Format last modified date
                    try:
                        if ticket.get("updated_at"):
                            last_modified = pd.to_datetime(
                                ticket["updated_at"]
                            ).strftime("%d/%m/%Y %H:%M")
                        else:
                            last_modified = "N/A"
                    except:
                        last_modified = ticket.get("updated_at", "N/A")

                    # Create a row for each specialty (or one row if no specialties)
                    if specialty_names:
                        for specialty_name in specialty_names:
                            # Apply specialty filter at the row level
                            if specialty_filter:
                                # Only include this row if the specialty matches the filter
                                if specialty_name not in specialty_filter:
                                    continue

                            enhanced_data.append(
                                {
                                    "Ticket ID": ticket["id"],
                                    "Customer Name": ticket["customer_name"],
                                    "Phone": ticket["customer_phone"],
                                    "Problem Description": (
                                        ticket["problem_desc"][:50] + "..."
                                        if len(str(ticket["problem_desc"])) > 50
                                        else ticket["problem_desc"]
                                    ),
                                    "Domain": domain_name,
                                    "Specialty": specialty_name,
                                    "Payment Status": (
                                        "Paid" if ticket["is_paid"] == 1 else "Unpaid"
                                    ),
                                    "Amount (₦)": (
                                        f"₦{ticket['amount']:,.0f}"
                                        if ticket["amount"] > 0
                                        else "₦0"
                                    ),
                                    "Created By": ticket.get(
                                        "created_by_name", "Unknown"
                                    ),
                                    "Creation Date": created_date,
                                    "Last Modified": last_modified,
                                }
                            )
                    else:
                        # If no specialties, create one row with "N/A" only if no specialty filter is applied
                        if not specialty_filter:
                            enhanced_data.append(
                                {
                                    "Ticket ID": ticket["id"],
                                    "Customer Name": ticket["customer_name"],
                                    "Phone": ticket["customer_phone"],
                                    "Problem Description": (
                                        ticket["problem_desc"][:50] + "..."
                                        if len(str(ticket["problem_desc"])) > 50
                                        else ticket["problem_desc"]
                                    ),
                                    "Domain": domain_name,
                                    "Specialty": "N/A",
                                    "Payment Status": (
                                        "Paid" if ticket["is_paid"] == 1 else "Unpaid"
                                    ),
                                    "Amount (₦)": (
                                        f"₦{ticket['amount']:,.0f}"
                                        if ticket["amount"] > 0
                                        else "₦0"
                                    ),
                                    "Created By": ticket.get(
                                        "created_by_name", "Unknown"
                                    ),
                                    "Creation Date": created_date,
                                    "Last Modified": last_modified,
                                }
                            )

                display_df = pd.DataFrame(enhanced_data)
                st.dataframe(display_df, use_container_width=True, hide_index=True)

                # Export functionality
                st.subheader("📤 Export Data")

                export_col1, export_col2, export_col3 = st.columns(3)

                with export_col1:
                    csv_data = export_to_csv(display_df)
                    if csv_data:
                        st.download_button(
                            label="📄 Export to CSV",
                            data=csv_data,
                            file_name=f"ticket_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                            mime="text/csv",
                            help="Download the filtered data as CSV file",
                        )

                with export_col2:
                    pdf_data = export_to_pdf(display_df, "Ticket Statistics Report")
                    if pdf_data:
                        st.download_button(
                            label="📄 Export to PDF",
                            data=pdf_data,
                            file_name=f"ticket_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                            mime="application/pdf",
                            help="Download the filtered data as PDF file",
                        )

                with export_col3:
                    excel_data = export_to_excel(display_df, "Ticket Statistics Report")
                    if excel_data:
                        st.download_button(
                            label="📊 Export to Excel",
                            data=excel_data,
                            file_name=f"ticket_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Download the filtered data as Excel file",
                        )



                    # Refresh button
                    if st.button("🔄 Refresh Statistics", key="refresh_stats"):
                        clear_cache()
                        st.rerun()
