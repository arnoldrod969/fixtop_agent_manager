
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import sys
import os
import random
import re
import time
import plotly.express as px
import plotly.graph_objects as go
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Add parent directory to path to import database and permissions
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database import db_manager
from permissions import PermissionManager

def export_to_csv(dataframe):
    """Export dataframe to CSV format"""
    try:
        return dataframe.to_csv(index=False).encode('utf-8')
    except Exception as e:
        st.error(f"Error exporting to CSV: {str(e)}")
        return None

def export_to_pdf(dataframe, title="Team Statistics Report"):
    """Export dataframe to PDF format"""
    try:
        buffer = io.BytesIO()
        # Utiliser des marges plus petites pour plus d'espace
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=20,
            alignment=1  # Center alignment
        )
        
        # Add title
        title_para = Paragraph(title, title_style)
        elements.append(title_para)
        elements.append(Spacer(1, 8))
        
        # Add generation date
        date_para = Paragraph(f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 12))
        
        # Prepare data for table avec troncature plus agressive
        data = [list(dataframe.columns)]  # Header
        for _, row in dataframe.iterrows():
            processed_row = []
            for i, cell in enumerate(row):
                cell_str = str(cell)
                # Troncature différente selon la colonne
                if i == 0:  # ID
                    processed_row.append(cell_str[:8])
                elif i == 1:  # Team Name
                    processed_row.append(cell_str[:15])
                elif i == 2:  # Code
                    processed_row.append(cell_str[:10])
                elif i == 3:  # Description
                    processed_row.append(cell_str[:20] + '...' if len(cell_str) > 20 else cell_str)
                elif i == 4:  # Manager
                    processed_row.append(cell_str[:15])
                elif i == 5:  # Member
                    processed_row.append(cell_str[:12])
                elif i == 6:  # Member Email
                    processed_row.append(cell_str[:20] + '...' if len(cell_str) > 20 else cell_str)
                elif i == 7:  # Member Role
                    processed_row.append(cell_str[:10])
                else:  # Dates
                    processed_row.append(cell_str[:12])
            data.append(processed_row)
        
        # Calculer la largeur disponible
        available_width = A4[0] - 60  # Largeur A4 moins les marges
        col_count = len(dataframe.columns)
        col_width = available_width / col_count
        
        # Créer le tableau avec largeurs de colonnes spécifiées
        table = Table(data, colWidths=[col_width] * col_count)
        
        # Style the table
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alignement à gauche pour économiser l'espace
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),  # Police plus petite
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),  # Police encore plus petite pour les données
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error exporting to PDF: {str(e)}")
        return None

def export_to_excel(dataframe, title="Team Statistics Report"):
    """Export dataframe to Excel format"""
    try:
        buffer = io.BytesIO()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Team Statistics"
        
        # Add title
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fusionner les cellules pour le titre en fonction du nombre de colonnes
        max_col = len(dataframe.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        
        # Add generation date
        ws['A2'] = f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        # Add data starting from row 4
        start_row = 4
        
        # Add headers manually
        for col_idx, column_name in enumerate(dataframe.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data rows
        for row_idx, (_, row) in enumerate(dataframe.iterrows(), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value))
        
        # Auto-adjust column widths - correction de l'erreur
        for col_idx in range(1, len(dataframe.columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Parcourir toutes les cellules de la colonne pour trouver la largeur max
            for row_idx in range(start_row, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_length = len(str(cell_value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            # Ajuster la largeur avec une limite max
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error exporting to Excel: {str(e)}")
        return None

def create_team_charts(teams_df):
    """Create charts for team visualization"""
    try:
        charts = {}
        
        if teams_df.empty:
            return charts
        
        # Chart 1: Teams by Manager
        manager_data = []
        for _, team in teams_df.iterrows():
            team_details = db_manager.get_team_by_id(team['id'])
            manager_name = team_details.get('manager_name', 'No manager') if team_details else 'No manager'
            manager_data.append(manager_name)
        
        if manager_data:
            manager_counts = pd.Series(manager_data).value_counts()
            charts['managers'] = px.pie(
                values=manager_counts.values,
                names=manager_counts.index,
                title="Teams by Manager"
            )
        
        # Chart 2: Team sizes distribution
        size_data = []
        for _, team in teams_df.iterrows():
            members = db_manager.get_team_members(team['id'])
            size_data.append(len(members))
        
        if size_data:
            size_df = pd.DataFrame({'Team Size': size_data})
            charts['sizes'] = px.histogram(
                size_df,
                x='Team Size',
                title="Team Size Distribution",
                nbins=10
            )
        
        # Chart 3: Teams creation timeline
        if 'created_at' in teams_df.columns:
            timeline_df = teams_df.copy()
            timeline_df['created_at'] = pd.to_datetime(timeline_df['created_at'])
            timeline_df['month'] = timeline_df['created_at'].dt.to_period('M')
            monthly_counts = timeline_df.groupby('month').size().reset_index(name='count')
            monthly_counts['month'] = monthly_counts['month'].astype(str)
            
            charts['timeline'] = px.line(
                monthly_counts,
                x='month',
                y='count',
                title="Teams Creation Timeline",
                markers=True
            )
        
        return charts
        
    except Exception as e:
        st.error(f"Error creating charts: {str(e)}")
        return {}

# Email validation function
def is_valid_email(email):
    """Validates email format"""
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(email_pattern, email) is not None

if 'logged_in' not in st.session_state or not st.session_state.logged_in:
    st.switch_page("app.py")  # Redirect to home/login if not connected

# Check page access permissions
if not PermissionManager.check_page_access('teams_page'):
    PermissionManager.show_access_denied("You do not have the necessary permissions to access this page.")

"""Team management page"""
st.title("🤖 Team Management")

# Define data loading functions
def load_teams_data():
    """Load teams data from database"""
    try:
        teams = db_manager.get_teams()
        if teams:
            df = pd.DataFrame(teams)
            # Convert dates
            if 'created_at' in df.columns:
                df['created_at'] = pd.to_datetime(df['created_at'])
            if 'updated_at' in df.columns:
                df['updated_at'] = pd.to_datetime(df['updated_at'])
            return df
        else:
            # Return empty DataFrame with expected columns
            return pd.DataFrame(columns=['id', 'name', 'description', 'created_at', 'created_by'])
    except Exception as e:
        st.error(f"Error loading teams: {str(e)}")
        return pd.DataFrame(columns=['id', 'name', 'description', 'created_at', 'created_by'])

# Get available tabs according to permissions
available_tabs = PermissionManager.get_available_tabs('teams_page')

if not available_tabs:
    st.error("No tabs available for your role.")
    st.stop()

# Create tabs dynamically
tabs = st.tabs(available_tabs)

# List Tab
if "📋 List" in available_tabs:
    tab_index = available_tabs.index("📋 List")
    with tabs[tab_index]:
        st.subheader("📋 Teams List")

        # Load teams data
        teams_df = load_teams_data()
        
        # Advanced Filters Section (moved from List tab)
        st.subheader("🔍 Advanced Filters")
        
        # Create filter columns
        filter_col1, filter_col2, filter_col3, filter_col4 = st.columns(4)

        with filter_col1:
            search_team = st.text_input("🔍 Search", placeholder="Name, description, code...", key="stats_search")

        with filter_col2:
            # Manager filter
            all_managers = []
            if not teams_df.empty:
                for _, team in teams_df.iterrows():
                    team_details = db_manager.get_team_by_id(team['id'])
                    if team_details and team_details.get('manager_name'):
                        all_managers.append(team_details['manager_name'])
                all_managers = sorted(list(set(all_managers)))
            
            manager_filter = st.selectbox(
                "👨‍💼 Filter by Manager",
                ["All managers"] + all_managers,
                help="Filter teams by assigned manager",
                key="stats_manager_filter"
            )

        with filter_col3:
            # Member count filter
            member_count_filter = st.selectbox(
                "👥 Number of members",
                ["All", "No members (0)", "Small teams (1-5)", "Medium teams (6-15)", "Large teams (16+)"],
                help="Filter by number of members",
                key="stats_member_filter"
            )

        with filter_col4:
            # Date filter
            date_filter = st.selectbox(
                "📅 Date Filter",
                ["All dates", "Creation date", "Modification date"],
                key="stats_date_filter"
            )
        
        # Date range inputs
        start_date = None
        end_date = None
        if date_filter != "All dates":
            date_col1, date_col2 = st.columns(2)
            with date_col1:
                start_date = st.date_input("From date", key="stats_start_date")
            with date_col2:
                end_date = st.date_input("To date", key="stats_end_date")
        
        # Apply advanced filters (same logic as before)
        filtered_df = teams_df.copy()
        if not filtered_df.empty:
            # Apply search filter
            if search_team.strip():
                search_term = search_team.strip().lower()
                team_codes = {}
                for _, team in filtered_df.iterrows():
                    team_details = db_manager.get_team_by_id(team['id'])
                    if team_details:
                        team_codes[team['id']] = team_details.get('code', '').lower()
                
                filtered_df = filtered_df[
                    filtered_df['name'].str.lower().str.contains(search_term, na=False) |
                    filtered_df['description'].str.lower().str.contains(search_term, na=False) |
                    filtered_df['id'].astype(str).str.contains(search_term, na=False) |
                    filtered_df['id'].apply(lambda x: search_term in team_codes.get(x, ''))
                ]
            
            # Apply manager filter
            if manager_filter != "All managers":
                team_ids_with_manager = []
                for _, team in filtered_df.iterrows():
                    team_details = db_manager.get_team_by_id(team['id'])
                    if team_details and team_details.get('manager_name') == manager_filter:
                        team_ids_with_manager.append(team['id'])
                filtered_df = filtered_df[filtered_df['id'].isin(team_ids_with_manager)]
            
            # Apply member count filter
            if member_count_filter != "All":
                team_ids_by_member_count = []
                for _, team in filtered_df.iterrows():
                    member_count = len(db_manager.get_team_members(team['id']))
                    if member_count_filter == "No members (0)" and member_count == 0:
                        team_ids_by_member_count.append(team['id'])
                    elif member_count_filter == "Small teams (1-5)" and 1 <= member_count <= 5:
                        team_ids_by_member_count.append(team['id'])
                    elif member_count_filter == "Medium teams (6-15)" and 6 <= member_count <= 15:
                        team_ids_by_member_count.append(team['id'])
                    elif member_count_filter == "Large teams (16+)" and member_count >= 16:
                        team_ids_by_member_count.append(team['id'])
                filtered_df = filtered_df[filtered_df['id'].isin(team_ids_by_member_count)]
            
            # Apply date filter
            if date_filter == "Creation date" and start_date is not None and end_date is not None:
                filtered_df['created_at_dt'] = pd.to_datetime(filtered_df['created_at'])
                filtered_df = filtered_df[
                    (filtered_df['created_at_dt'].dt.date >= start_date) &
                    (filtered_df['created_at_dt'].dt.date <= end_date)
                ]
                filtered_df = filtered_df.drop('created_at_dt', axis=1)
            elif date_filter == "Modification date" and start_date is not None and end_date is not None:
                if 'updated_at' in filtered_df.columns:
                    filtered_df['updated_at_dt'] = pd.to_datetime(filtered_df['updated_at'])
                    filtered_df = filtered_df[
                        (filtered_df['updated_at_dt'].dt.date >= start_date) &
                        (filtered_df['updated_at_dt'].dt.date <= end_date)
                    ]
                    filtered_df = filtered_df.drop('updated_at_dt', axis=1)

        # Display statistics and charts
        if not filtered_df.empty:
            # Enhanced data table with metrics - One row per member
            # st.subheader("📋 Detailed Team List with Members")
            
            enhanced_data = []
            for _, team in filtered_df.iterrows():
                members = db_manager.get_team_members(team['id'])
                
                try:
                    created_date = pd.to_datetime(team['created_at']).strftime('%d/%m/%Y')
                except:
                    created_date = team['created_at']
                
                # Use manager_name directly from teams_df (already available from get_teams())
                manager_name = team.get('manager_name', 'Not assigned') or 'Not assigned'
                
                # If team has no members, show one row with "No members"
                if not members:
                    enhanced_data.append({
                        'Team ID': team['id'],
                        'Team Name': team['name'],
                        'Code': team.get('code', 'N/A'),
                        'Description': team['description'][:50] + '...' if len(str(team['description'])) > 50 else team['description'],
                        'Manager': manager_name,
                        'Member': 'No members',
                        'Member Email': '-',
                        'Member Role': '-',
                        'Creation date': created_date,
                        'Last modified': team.get('updated_at', 'N/A')
                    })
                else:
                    # Create one row per member
                    for member in members:
                        enhanced_data.append({
                            'Team ID': team['id'],
                            'Team Name': team['name'],
                            'Code': team.get('code', 'N/A'),
                            'Description': team['description'][:50] + '...' if len(str(team['description'])) > 50 else team['description'],
                            'Manager': manager_name,
                            'Member': member.get('user_name', 'Unknown'),
                            'Member Email': member.get('user_email', 'N/A'),
                            'Member Role': member.get('user_role', 'N/A'),
                            'Creation date': created_date,
                            'Last modified': team.get('updated_at', 'N/A')
                        })
            
            display_df = pd.DataFrame(enhanced_data)
            st.dataframe(display_df, use_container_width=True, hide_index=True)
            
            # Export functionality
            st.subheader("📤 Export Data")
            
            export_col1, export_col2, export_col3 = st.columns(3)
            
            with export_col1:
                if st.button("📊 Export to CSV", key="export_csv"):
                    csv_data = export_to_csv(display_df)
                    st.download_button(
                        label="⬇️ Download CSV",
                        data=csv_data,
                        file_name=f"teams_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
                    st.success("✅ CSV export ready for download!")
            
            with export_col2:
                if st.button("📋 Export Filtered Data", key="export_filtered"):
                    st.info("📋 Filtered data export prepared")
            
            with export_col3:
                if st.button("📈 Export Charts", key="export_charts"):
                    st.info("📈 Chart export feature coming soon")
        
        else:
            st.info("🔍 No teams found with the applied filters.")
            st.write("💡 **Suggestions:**")
            st.write("• Adjust your filter criteria")
            st.write("• Try a broader date range")
            st.write("• Check if teams exist in the system")

# Add Tab
if "➕ Add" in available_tabs:
    tab_index = available_tabs.index("➕ Add")
    with tabs[tab_index]:
        st.subheader("➕ Add Team")
        
        # Get available managers (users with manager role who are not already managing a team)
        def get_available_managers():
            """Get managers who are not already assigned to a team"""
            try:
                all_users = db_manager.get_all_users()
                if all_users:
                    # Filter for active managers
                    managers = [user for user in all_users if user['role_name'] == 'manager' and user['is_active'] == 1]
                    
                    # Check which managers are not already assigned to a team
                    available_managers = []
                    for manager in managers:
                        if db_manager.is_manager_available(manager['id']):
                            available_managers.append(manager)
                    
                    return available_managers
                return []
            except Exception as e:
                st.error(f"Error loading available managers: {str(e)}")
                return []
        
        available_managers = get_available_managers()
        
        if not available_managers:
            st.warning("⚠️ No available managers found. All managers are already assigned to teams or no managers exist.")
            st.info("💡 You need to create managers first or free up existing managers from their current teams.")
        else:
            with st.form("add_team_form"):
                col1, col2 = st.columns(2)
                
                with col1:
                    # Team name with real-time validation hint
                    name = st.text_input("Team Name *", placeholder="Ex: Technical Support Team")
                    if name.strip():
                        # Check if name already exists
                        existing_teams = db_manager.get_teams()
                        existing_names = [team['name'].lower() for team in existing_teams]
                        if name.strip().lower() in existing_names:
                            st.error("❌ This team name already exists. Please choose a different name.")
                    
                    # Manager selection dropdown
                    manager_options = {f"{manager['name']} ({manager['email']})": manager['id'] for manager in available_managers}
                    selected_manager_display = st.selectbox(
                        "Manager *", 
                        options=list(manager_options.keys()),
                        help="Select the manager who will be responsible for this team"
                    )
                    # Initialiser selected_manager_id avec une valeur par défaut
                    selected_manager_id = None
                    if selected_manager_display and selected_manager_display in manager_options:
                        selected_manager_id = manager_options[selected_manager_display]
                    
                    # Description (optional)
                    description = st.text_area("Description", placeholder="Team description and responsibilities...")
                    
                    # Team members selection (optional)
                    st.write("**Team Members (Optional)**")
                    available_users = db_manager.get_all_users()
                    if available_users:
                        # Filter out inactive users, the selected manager, and non-agent users
                        available_members = [
                            user for user in available_users 
                            if user['is_active'] == 1 
                            and (selected_manager_id is None or user['id'] != selected_manager_id)
                            and user['role_name'].lower() == 'agent'
                        ]
                        
                        if available_members:
                            member_options = [f"{user['name']} - {user['email']} ({user['role_name']})" for user in available_members]
                            selected_members = st.multiselect(
                                "Select Team Members",
                                options=member_options,
                                help="Select users to add as team members (optional)"
                            )
                            
                            # Get selected member IDs
                            selected_member_ids = []
                            for selected_member in selected_members:
                                for user in available_members:
                                    if f"{user['name']} - {user['email']} ({user['role_name']})" == selected_member:
                                        selected_member_ids.append(user['id'])
                                        break
                        else:
                            selected_member_ids = []
                            st.info("No available users to add as members")
                    else:
                        selected_member_ids = []
                        st.info("No users available")
                
                with col2:
                    st.write("**Team Information**")
                    st.info("📋 The team name must be unique")
                    st.info("👤 Each manager can only manage one team")
                    st.info("🔢 Team code will be generated automatically")
                    st.info("👥 You can add members during team creation")
                    
                    # Show selected manager info
                    if selected_manager_display:
                        selected_manager = next(m for m in available_managers if m['id'] == selected_manager_id)
                        st.write("**Selected Manager:**")
                        st.write(f"• Name: {selected_manager['name']}")
                        st.write(f"• Email: {selected_manager['email']}")
                    
                    # Show selected members info
                    if 'selected_member_ids' in locals() and selected_member_ids:
                        st.write("**Selected Members:**")
                        for member_id in selected_member_ids:
                            member = next((u for u in available_members if u['id'] == member_id), None)
                            if member:
                                st.write(f"• {member['name']} ({member['role_name']})")
                
                submitted = st.form_submit_button("Create Team", type="primary")
                
                if submitted:
                    # Validation
                    errors = []
                    
                    if not name.strip():
                        errors.append("Team name is required")
                    
                    if not selected_manager_id:
                        errors.append("Manager selection is required")
                    
                    # Check team name uniqueness
                    if name.strip():
                        existing_teams = db_manager.get_teams()
                        existing_names = [team['name'].lower() for team in existing_teams]
                        if name.strip().lower() in existing_names:
                            errors.append("Team name already exists")
                    
                    # Check manager availability (double-check)
                    if selected_manager_id and not db_manager.is_manager_available(selected_manager_id):
                        errors.append("Selected manager is no longer available")
                    
                    if errors:
                        for error in errors:
                            st.error(f"❌ {error}")
                    else:
                        # Create team with manager_id
                        created_by = st.session_state.get('user_id', 1)
                        success, message, team_id = db_manager.create_team(
                            name=name.strip(),
                            description=description.strip() if description.strip() else None,
                            manager_id=selected_manager_id,
                            created_by=created_by
                        )
                        
                        if success:
                            st.success(f"✅ {message}")
                            # Show team code if available
                            if team_id:
                                team_info = db_manager.get_team_by_id(team_id)
                                if team_info and 'code' in team_info:
                                    st.info(f"🔢 Team code: **{team_info['code']}**")
                            
                            # Add selected members to the team
                            if selected_member_ids:
                                members_added = 0
                                members_failed = 0
                                for member_id in selected_member_ids:
                                    member_success, member_message = db_manager.add_team_member(
                                        team_id,
                                        member_id,
                                        created_by
                                    )
                                    if member_success:
                                        members_added += 1
                                    else:
                                        members_failed += 1
                                
                                if members_added > 0:
                                    st.success(f"✅ {members_added} member(s) added to the team successfully!")
                                if members_failed > 0:
                                    st.warning(f"⚠️ {members_failed} member(s) could not be added to the team.")
                            
                            st.balloons()
                            time.sleep(2)
                            st.rerun()
                        else:
                            st.error(f"❌ {message}")

# Delete Tab
if "🗑️ Delete" in available_tabs:
    tab_index = available_tabs.index("🗑️ Delete")
    with tabs[tab_index]:
        st.subheader("🗑️ Delete Team")
        st.warning("⚠️ **Attention:** Team deletion is irreversible!")
        
        teams_df = load_teams_data()
        
        if not teams_df.empty:
            # Team selection section
            st.subheader("🎯 Select Team to Delete")
            
            team_options = teams_df.apply(lambda x: f"{x['name']} - ID={x['id']}", axis=1).tolist()
            selected_option = st.selectbox(
                "Choose a team to delete",
                [""] + team_options,
                help="Select the team you want to delete"
            )
            
            if selected_option:
                selected_team_id = int(selected_option.split("ID=")[-1])
                team_data = teams_df[teams_df['id'] == selected_team_id].iloc[0]
                
                # Get detailed team information
                current_team = db_manager.get_team_by_id(selected_team_id)
                members = db_manager.get_team_members(selected_team_id)
                
                # Display team information
                with st.expander("📋 Team Information", expanded=True):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Name:** {team_data['name']}")
                        st.write(f"**Description:** {team_data.get('description', 'No description')}")
                        st.write(f"**Code:** {current_team.get('code', 'N/A') if current_team else 'N/A'}")
                    with col2:
                        manager_name = current_team.get('manager_name', 'No manager assigned') if current_team else 'No manager assigned'
                        st.write(f"**Manager:** {manager_name}")
                        st.write(f"**Number of members:** {len(members)}")
                        st.write(f"**Created on:** {team_data.get('created_at', 'Not available')}")
                
                # Members validation
                if members:
                    st.error("❌ **Cannot delete this team**")
                    st.write("This team still contains members. You must first remove all members before you can delete the team.")
                    
                    # Show current members
                    st.subheader("👥 Current Members")
                    for member in members:
                        st.write(f"• **{member['user_name']}** ({member['user_role']}) - {member['user_email']}")
                    
                    st.info("💡 **Tip:** Use the 'Edit' tab to remove members from this team before deleting it.")
                
                else:
                    # Team can be deleted
                    st.success("✅ **This team can be deleted**")
                    st.write("This team contains no members and can be safely deleted.")
                    
                    # Confirmation section
                    st.subheader("⚠️ Deletion Confirmation")
                    
                    confirmation_text = st.text_input(
                        f"Type the team name '{team_data['name']}' to confirm:",
                        placeholder=f"Type '{team_data['name']}' here"
                    )
                    
                    name_confirmed = confirmation_text.strip() == team_data['name']
                    
                    if not name_confirmed and confirmation_text.strip():
                        st.error("❌ Name does not match")
                    
                    # Final confirmation checkbox
                    final_confirmation = st.checkbox(
                        f"I confirm I want to delete the team '{team_data['name']}'",
                        key="final_delete_confirmation"
                    )
                    
                    # Delete button
                    if name_confirmed and final_confirmation:
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            if st.button("🗑️ Delete Team", type="primary", key="execute_delete_btn"):
                                success, message = db_manager.delete_team(
                                    selected_team_id,
                                    st.session_state.get('user_id', 1)
                                )
                                
                                if success:
                                    st.success(f"✅ {message}")
                                    st.balloons()
                                    time.sleep(2)
                                    st.rerun()
                                else:
                                    st.error(f"❌ {message}")
                        
                        with col2:
                            if st.button("❌ Cancel", key="cancel_delete_btn"):
                                st.session_state.final_delete_confirmation = False
                                st.rerun()
                    
                    elif not name_confirmed:
                        st.info("ℹ️ Please confirm the team name to continue")
                    elif not final_confirmation:
                        st.info("ℹ️ Please check the confirmation box to continue")
        
        else:
            st.info("No teams available for deletion.")

# Edit Tab
if "✏️ Edit" in available_tabs:
    tab_index = available_tabs.index("✏️ Edit")
    with tabs[tab_index]:
        st.subheader("✏️ Edit Team")
        teams_df = load_teams_data()
        
        if not teams_df.empty:
            # Team selection with improved UI
            st.subheader("🎯 Select Team to Edit")
            team_options = teams_df.apply(lambda x: f"{x['name']} - ID={x['id']}", axis=1).tolist()
            selected_option = st.selectbox(
                "Choose a team to edit",
                team_options,
                help="Select the team you want to modify"
            )
            
            try:
                selected_team_id = int(selected_option.split("ID=")[-1])
                team_data = teams_df[teams_df['id'] == selected_team_id].iloc[0]
                
                # Get current team details including manager
                current_team = db_manager.get_team_by_id(selected_team_id)
                current_manager_id = current_team.get('manager_id') if current_team else None
                current_manager_name = current_team.get('manager_name', 'No manager assigned') if current_team else 'No manager assigned'
            except (ValueError, IndexError) as e:
                st.error("❌ Error loading team data. Please try again.")
                st.stop()
            
            # Display current information
            with st.expander("📋 Current Information", expanded=True):
                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**Name:** {team_data['name']}")
                    st.write(f"**Description:** {team_data.get('description', 'No description')}")
                    st.write(f"**Code:** {current_team.get('code', 'N/A') if current_team else 'N/A'}")
                with col2:
                    st.write(f"**Current Manager:** {current_manager_name}")
                    members = db_manager.get_team_members(selected_team_id)
                    st.write(f"**Members:** {len(members)}")
                    try:
                        created_date = pd.to_datetime(team_data.get('created_at')).strftime('%d/%m/%Y %H:%M')
                    except:
                        created_date = team_data.get('created_at', 'Not available')
                    st.write(f"**Created on:** {created_date}")
            
            # Team Information Update Form
            with st.form(f"edit_team_form_{selected_team_id}"):
                st.subheader("🔧 Update Team Information")
                
                col1, col2 = st.columns(2)
                with col1:
                    new_name = st.text_input(
                        "Team Name *", 
                        value=team_data['name'] or "",
                        help="Enter a unique team name"
                    )
                    new_description = st.text_area(
                        "Description", 
                        value=team_data['description'] or "",
                        help="Describe the team's purpose and responsibilities"
                    )
                
                with col2:
                    # Manager selection
                    st.write("**Manager Assignment**")
                    try:
                        available_managers = get_available_managers()
                    except Exception as e:
                        st.error(f"❌ Error loading managers: {str(e)}")
                        available_managers = []
                    
                    # Add current manager to options if not in available list
                    manager_options = []
                    manager_ids = []
                    
                    if current_manager_id and current_manager_name != 'No manager assigned':
                        manager_options.append(f"{current_manager_name} (Current)")
                        manager_ids.append(current_manager_id)
                    
                    for manager in available_managers:
                        if manager['id'] != current_manager_id:
                            manager_options.append(f"{manager['name']} - {manager['email']}")
                            manager_ids.append(manager['id'])
                    
                    if manager_options:
                        selected_manager_index = st.selectbox(
                            "Select Manager",
                            range(len(manager_options)),
                            format_func=lambda x: manager_options[x],
                            index=0 if current_manager_id else None,
                            help="Choose a manager for this team"
                        )
                        selected_manager_id = manager_ids[selected_manager_index] if selected_manager_index is not None else None
                    else:
                        st.warning("⚠️ No available managers found")
                        selected_manager_id = current_manager_id
                    
                    st.info("💡 The team name must be unique")
                
                col1, col2 = st.columns(2)
                with col1:
                    update_submitted = st.form_submit_button("💾 Update Team", type="primary")
                with col2:
                    if st.form_submit_button("🔄 Reset"):
                        st.rerun()
                
                if update_submitted:
                    # Validation
                    errors = []
                    
                    if not new_name.strip():
                        errors.append("Team name is required")
                    
                    # Check team name uniqueness (only if name changed)
                    if new_name.strip() and new_name.strip().lower() != team_data['name'].lower():
                        try:
                            existing_teams = db_manager.get_teams()
                            existing_names = [team['name'].lower() for team in existing_teams if team['id'] != selected_team_id]
                            if new_name.strip().lower() in existing_names:
                                errors.append("Team name already exists")
                        except Exception as e:
                            errors.append(f"Error checking team name uniqueness: {str(e)}")
                    
                    # Check manager availability (only if manager changed)
                    if selected_manager_id and selected_manager_id != current_manager_id:
                        try:
                            if not db_manager.is_manager_available(selected_manager_id):
                                errors.append("Selected manager is no longer available")
                        except Exception as e:
                            errors.append(f"Error checking manager availability: {str(e)}")
                    
                    if errors:
                        for error in errors:
                            st.error(f"❌ {error}")
                    else:
                        # Prepare update parameters
                        update_params = {}
                        
                        # Only update fields that have changed
                        if new_name.strip() != team_data['name']:
                            update_params['name'] = new_name.strip()
                        
                        if new_description.strip() != (team_data['description'] or ""):
                            update_params['description'] = new_description.strip() if new_description.strip() else None
                        
                        if selected_manager_id != current_manager_id:
                            update_params['manager_id'] = selected_manager_id
                        
                        if update_params:
                            try:
                                updated_by = st.session_state.get('user_id', 1)
                                success, message = db_manager.update_team(
                                    team_id=selected_team_id,
                                    updated_by=updated_by,
                                    **update_params
                                )
                                if success:
                                    st.success(f"✅ {message}")
                                    st.balloons()
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.error(f"❌ {message}")
                            except Exception as e:
                                st.error(f"❌ Error updating team: {str(e)}")
                        else:
                            st.info("ℹ️ No changes detected")
            
            # Team Members Management Section
            st.divider()
            st.subheader("👥 Team Members Management")
            
            # Current members
            try:
                members = db_manager.get_team_members(selected_team_id)
            except Exception as e:
                st.error(f"❌ Error loading team members: {str(e)}")
                members = []
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Current Members**")
                if members:
                    for member in members:
                        with st.container():
                            member_col1, member_col2 = st.columns([3, 1])
                            with member_col1:
                                st.write(f"• **{member['user_name']}** ({member['user_role']})")
                                st.caption(f"📧 {member['user_email']}")
                            with member_col2:
                                if st.button("🗑️", key=f"remove_member_{member['user_id']}", help="Remove member"):
                                    try:
                                        success, message = db_manager.remove_team_member(
                                            selected_team_id,
                                            member['user_id'],
                                            st.session_state.get('user_id', 1)
                                        )
                                        if success:
                                            st.success(f"✅ {message}")
                                            time.sleep(1)
                                            st.rerun()
                                        else:
                                            st.error(f"❌ {message}")
                                    except Exception as e:
                                        st.error(f"❌ Error removing member: {str(e)}")
                else:
                    st.info("👤 No members in this team")
                    st.write("💡 **Tip:** Use the 'Add New Member' section to add team members.")
            
            with col2:
                st.write("**Add New Member**")
                try:
                    available_users = db_manager.get_available_users_for_team(selected_team_id)
                except Exception as e:
                    st.error(f"❌ Error loading available users: {str(e)}")
                    available_users = []
                
                if available_users:
                    user_options = [f"{user['name']} - {user['email']} (agent)" for user in available_users]
                    selected_user_index = st.selectbox(
                        "Select Agent to Add",
                        range(len(user_options)),
                        format_func=lambda x: user_options[x],
                        key="add_member_select",
                        help="Choose an agent to add to this team"
                    )
                    
                    if st.button("➕ Add Member", key="add_member_btn", type="primary"):
                        try:
                            selected_user = available_users[selected_user_index]
                            
                            # Validation supplémentaire : vérifier que l'utilisateur est un agent disponible
                            if selected_user['role'].lower() != 'agent':
                                st.error(f"❌ Only agents can be added as team members. Selected user has role: {selected_user['role']}")
                            elif not db_manager.is_agent_available(selected_user['id']):
                                st.error(f"❌ This agent is already a member of another active team")
                            else:
                                success, message = db_manager.add_team_member(
                                    selected_team_id,
                                    selected_user['id'],
                                    st.session_state.get('user_id', 1)
                                )
                                if success:
                                    st.success(f"✅ {message}")
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.error(f"❌ {message}")
                        except Exception as e:
                            st.error(f"❌ Error adding member: {str(e)}")
                else:
                    st.info("👥 No available agents to add")
                    st.write("💡 **Suggestions:**")
                    st.write("• All eligible agents are already team members")
                    st.write("• Create new agents first")
                    st.write("• Check that agents are not already assigned to other teams")
        
        else:
            st.info("🔍 No teams found.")
            st.write("💡 **Suggestion:** Create teams first using the '➕ Add' tab.")
            st.write("📋 You need to have at least one team to use the editing features.")

# Statistics Tab
if "📊 Statistics" in available_tabs:
    tab_index = available_tabs.index("📊 Statistics")
    with tabs[tab_index]:
        st.subheader("📊 Team Statistics")
        
        # Load teams data
        teams_df = load_teams_data()
        
        if not teams_df.empty:
            # Filters Section
            st.subheader("🔍 Filters")
            
            # Create filter columns
            filter_col1, filter_col2, filter_col3, filter_col4 = st.columns(4)
            
            with filter_col1:
                search_team = st.text_input("🔍 Search Team", placeholder="Team name or ID...", key="stats_search_team")
            
            with filter_col2:
                # Get all active managers from database (not just those assigned to teams)
                try:
                    all_users = db_manager.get_all_users()
                    all_managers = []
                    if all_users:
                        # Filter for active managers
                        active_managers = [user for user in all_users if user['role_name'] == 'manager' and user['is_active'] == 1]
                        all_managers = [manager['name'] for manager in active_managers]
                    unique_managers = sorted(list(set(all_managers)))
                    selected_managers = st.multiselect(
                        "👤 Filter by Manager(s)", 
                        options=unique_managers,
                        help="Select one or more managers to filter teams",
                        key="statistics_manager_filter"
                    )
                except Exception as e:
                    st.error(f"Error loading managers: {str(e)}")
                    selected_managers = []
            
            with filter_col3:
                # Get all active agents from database (not just those assigned to teams)
                try:
                    all_users = db_manager.get_all_users()
                    all_agents = []
                    if all_users:
                        # Filter for active agents
                        active_agents = [user for user in all_users if user['role_name'] == 'agent' and user['is_active'] == 1]
                        all_agents = [agent['name'] for agent in active_agents]
                    unique_agents = sorted(list(set(all_agents)))
                    selected_agents = st.multiselect(
                        "🧑‍💼 Filter by Agent(s)", 
                        options=unique_agents,
                        help="Select one or more agents to filter teams",
                        key="statistics_agent_filter"
                    )
                except Exception as e:
                    st.error(f"Error loading agents: {str(e)}")
                    selected_agents = []
            
            with filter_col4:
                # Date filters
                date_filter_type = st.selectbox("📅 Date Filter", ["All", "Creation Date", "Modification Date"], key="statistics_date_filter")
            
            # Date range inputs
            start_date = None
            end_date = None
            if date_filter_type != "All":
                date_col1, date_col2 = st.columns(2)
                
                # Calculer des valeurs par défaut intelligentes
                if not teams_df.empty:
                    min_date = teams_df['created_at'].min().date() if date_filter_type == "Creation Date" else teams_df['updated_at'].min().date() if 'updated_at' in teams_df.columns else teams_df['created_at'].min().date()
                    max_date = teams_df['created_at'].max().date() if date_filter_type == "Creation Date" else teams_df['updated_at'].max().date() if 'updated_at' in teams_df.columns else teams_df['created_at'].max().date()
                else:
                    from datetime import date, timedelta
                    min_date = date.today() - timedelta(days=30)
                    max_date = date.today()
                
                with date_col1:
                    start_date = st.date_input("From", value=min_date, key="statistics_start_date")
                with date_col2:
                    end_date = st.date_input("To", value=max_date, key="statistics_end_date")
            
            # Apply filters
            filtered_teams = teams_df.copy()
            
            # Search filter
            if search_team:
                filtered_teams = filtered_teams[
                    filtered_teams['name'].str.contains(search_team, case=False, na=False) |
                    filtered_teams['id'].astype(str).str.contains(search_team, case=False, na=False)
                ]
            
            # Manager filter
            if selected_managers:  # If managers are selected
                team_ids_with_manager = []
                for _, team in filtered_teams.iterrows():
                    if team.get('manager_name') in selected_managers:
                        team_ids_with_manager.append(team['id'])
                filtered_teams = filtered_teams[filtered_teams['id'].isin(team_ids_with_manager)]
            
            # Agent filter
            if selected_agents:
                team_ids_with_agent = []
                for _, team in filtered_teams.iterrows():
                    members = db_manager.get_team_members(team['id'])
                    for member in members:
                        if member.get('user_name') in selected_agents:
                            team_ids_with_agent.append(team['id'])
                            break
                filtered_teams = filtered_teams[filtered_teams['id'].isin(team_ids_with_agent)]
            
            # Date filter
            if date_filter_type == "Creation Date" and start_date is not None and end_date is not None:
                filtered_teams = filtered_teams[
                    (filtered_teams['created_at'].dt.date >= start_date) &
                    (filtered_teams['created_at'].dt.date <= end_date)
                ]
            elif date_filter_type == "Modification Date" and start_date is not None and end_date is not None:
                if 'updated_at' in filtered_teams.columns:
                    filtered_teams = filtered_teams[
                        (filtered_teams['updated_at'].dt.date >= start_date) &
                        (filtered_teams['updated_at'].dt.date <= end_date)
                    ]
            
            # Key Metrics
            st.subheader("📈 Key Metrics")
            
            # Calculate metrics
            total_teams = len(filtered_teams)
            total_members = 0
            teams_with_managers = 0
            empty_teams = 0
            
            for _, team in filtered_teams.iterrows():
                team_details = db_manager.get_team_by_id(team['id'])
                if team_details:
                    members_count = len(team_details.get('members', []))
                    total_members += members_count
                    if team_details.get('manager_name'):
                        teams_with_managers += 1
                    if members_count == 0:
                        empty_teams += 1
            
            avg_team_size = total_members / total_teams if total_teams > 0 else 0
            
            # Key metrics
            col1, col2, col3, col4 = st.columns(4)
            
            total_teams = len(filtered_teams)
            total_members = sum(len(db_manager.get_team_members(team['id'])) for _, team in filtered_teams.iterrows())
            avg_team_size = total_members / total_teams if total_teams > 0 else 0
            teams_without_members = sum(1 for _, team in filtered_teams.iterrows() if len(db_manager.get_team_members(team['id'])) == 0)
            
            with col1:
                st.metric("📊 Total Teams", total_teams)
            with col2:
                st.metric("👥 Total Members", total_members)
            with col3:
                st.metric("📈 Avg Team Size", f"{avg_team_size:.1f}")
            with col4:
                st.metric("⚠️ Empty Teams", teams_without_members)
            
            # Simple data table - One row per member (similar to List tab)
            # st.subheader("📋 Detailed Team List with Members")
            
            enhanced_data = []
            for _, team in filtered_teams.iterrows():
                members = db_manager.get_team_members(team['id'])
                
                try:
                    created_date = pd.to_datetime(team['created_at']).strftime('%d/%m/%Y')
                except:
                    created_date = team['created_at']
                
                # Use manager_name directly from teams_df (already available from get_teams())
                manager_name = team.get('manager_name', 'Not assigned') or 'Not assigned'
                
                # Get team details to retrieve the code
                team_details = db_manager.get_team_by_id(team['id'])
                team_code = team_details.get('code', 'N/A') if team_details else 'N/A'
                
                # If team has no members, show one row with "No members"
                if not members:
                    enhanced_data.append({
                        'Team ID': team['id'],
                        'Team Name': team['name'],
                        'Code': team_code,
                        'Description': team['description'][:50] + '...' if len(str(team['description'])) > 50 else team['description'],
                        'Manager': manager_name,
                        'Member': 'No members',
                        'Member Email': '-',
                        'Member Role': '-',
                        'Creation date': created_date,
                        'Last modified': team.get('updated_at', 'N/A')
                    })
                else:
                    # Create one row per member, but filter by selected agents if any
                    for member in members:
                        # If agents are selected, only show members that match the selected agents
                        if selected_agents and member.get('user_name') not in selected_agents:
                            continue
                            
                        enhanced_data.append({
                            'Team ID': team['id'],
                            'Team Name': team['name'],
                            'Code': team_code,
                            'Description': team['description'][:50] + '...' if len(str(team['description'])) > 50 else team['description'],
                            'Manager': manager_name,
                            'Member': member.get('user_name', 'Unknown'),
                            'Member Email': member.get('user_email', 'N/A'),
                            'Member Role': member.get('user_role', 'N/A'),
                            'Creation date': created_date,
                            'Last modified': team.get('updated_at', 'N/A')
                        })
            
            display_df = pd.DataFrame(enhanced_data)
            st.dataframe(display_df, use_container_width=True, hide_index=True)
            
            # Export functionality
            st.subheader("📤 Export Data")
            
            export_col1, export_col2, export_col3 = st.columns(3)
            
            with export_col1:
                csv_data = export_to_csv(display_df)
                if csv_data:
                    st.download_button(
                        label="📄 Export to CSV",
                        data=csv_data,
                        file_name=f"team_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                        help="Download the filtered data as CSV file"
                    )
            
            with export_col2:
                pdf_data = export_to_pdf(display_df, "Team Statistics Report")
                if pdf_data:
                    st.download_button(
                        label="📄 Export to PDF",
                        data=pdf_data,
                        file_name=f"team_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf",
                        help="Download the filtered data as PDF file"
                    )
            
            with export_col3:
                excel_data = export_to_excel(display_df, "Team Statistics Report")
                if excel_data:
                    st.download_button(
                        label="📊 Export to Excel",
                        data=excel_data,
                        file_name=f"team_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Download the filtered data as Excel file"
                    )
            
            # Charts Section in expandable container
            with st.expander("📊 Charts & Analytics", expanded=False):
                if not filtered_teams.empty:
                    chart_col1, chart_col2 = st.columns(2)
                    
                    with chart_col1:
                        # Chart 1: Teams by Manager (Bar Chart)
                        manager_data = []
                        for _, team in filtered_teams.iterrows():
                            team_details = db_manager.get_team_by_id(team['id'])
                            manager_name = team_details.get('manager_name', 'No Manager') if team_details else 'No Manager'
                            manager_data.append(manager_name)
                        
                        if manager_data:
                            manager_counts = pd.Series(manager_data).value_counts()
                            fig_bar = px.bar(
                                x=manager_counts.index,
                                y=manager_counts.values,
                                title="Number of Teams by Manager",
                                labels={'x': 'Manager', 'y': 'Number of Teams'}
                            )
                            fig_bar.update_layout(height=400)
                            st.plotly_chart(fig_bar, use_container_width=True)
                    
                    with chart_col2:
                        # Chart 2: Agents Distribution by Team (Pie Chart)
                        team_sizes = []
                        team_names = []
                        for _, team in filtered_teams.iterrows():
                            team_details = db_manager.get_team_by_id(team['id'])
                            if team_details:
                                members_count = len(team_details.get('members', []))
                                if members_count > 0:  # Only show teams with members
                                    team_sizes.append(members_count)
                                    team_names.append(team['name'])
                        
                        if team_sizes:
                            fig_pie = px.pie(
                                values=team_sizes,
                                names=team_names,
                                title="Agent Distribution by Team"
                            )
                            fig_pie.update_layout(height=400)
                            st.plotly_chart(fig_pie, use_container_width=True)
                    
                    # Chart 3: Team Evolution Over Time (Line Chart)
                    st.subheader("📈 Team Evolution Over Time")
                    
                    if 'created_at' in filtered_teams.columns:
                        # Group teams by creation date
                        teams_by_date = filtered_teams.groupby(filtered_teams['created_at'].dt.date).size().cumsum()
                        
                        fig_line = px.line(
                            x=teams_by_date.index,
                            y=teams_by_date.values,
                            title="Cumulative Number of Teams Over Time",
                            labels={'x': 'Date', 'y': 'Total Teams'}
                        )
                        fig_line.update_layout(height=400)
                        st.plotly_chart(fig_line, use_container_width=True)
                else:
                    st.info("📊 No data available for charts. Create teams first to view analytics.")
        
        else:
            st.info("📊 No teams found. Create teams first to view statistics.")
            st.write("💡 **Suggestion:** Use the '➕ Add' tab to create your first team.")