import io
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def export_to_csv(dataframe):
    """Export dataframe to CSV format"""
    try:
        return dataframe.to_csv(index=False).encode("utf-8")
    except Exception as e:
        st.error(f"Error exporting to CSV: {str(e)}")
        return None



def export_to_pdf(data, title="Ticket Statistics Report"):
    """Export data to PDF format
    Args:
        data: Either a DataFrame or a dictionary of DataFrames
        title: Report title
    """
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=50,
            bottomMargin=30,
        )

        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=14,
            spaceAfter=20,
            alignment=1,
        )

        title_para = Paragraph(title, title_style)
        elements.append(title_para)
        elements.append(Spacer(1, 8))

        date_para = Paragraph(
            f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles["Normal"],
        )
        elements.append(date_para)
        elements.append(Spacer(1, 12))

        # Handle both single DataFrame and dictionary of DataFrames
        if isinstance(data, dict):
            dataframes = data
        else:
            dataframes = {"Data": data}

        for section_name, dataframe in dataframes.items():
            if dataframe.empty:
                continue
                
            # Add section title if multiple sections
            if len(dataframes) > 1:
                section_title = Paragraph(f"{section_name}", styles["Heading2"])
                elements.append(section_title)
                elements.append(Spacer(1, 6))

            # Prepare table data
            table_data = [list(dataframe.columns)]
            for _, row in dataframe.iterrows():
                processed_row = []
                for i, cell in enumerate(row):
                    cell_str = str(cell)
                    if i == 0:  # First column (usually ID)
                        processed_row.append(cell_str[:8])
                    elif i == 1:  # Second column (usually Name)
                        processed_row.append(cell_str[:15])
                    elif i == 2:  # Third column
                        processed_row.append(cell_str[:12])
                    elif i == 3:  # Fourth column (usually Description)
                        processed_row.append(
                            cell_str[:25] + "..." if len(cell_str) > 25 else cell_str
                        )
                    else:
                        processed_row.append(cell_str[:15])
                table_data.append(processed_row)

            available_width = A4[0] - 60
            col_count = len(dataframe.columns)
            col_width = available_width / col_count

            table = Table(table_data, colWidths=[col_width] * col_count)

            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 0), (-1, 0), 8),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 7),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 3),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )

            elements.append(table)
            
            # Add space between sections
            if len(dataframes) > 1:
                elements.append(Spacer(1, 12))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    except Exception as e:
        st.error(f"Error exporting to PDF: {str(e)}")
        return None


def export_to_excel(data, title="Ticket Statistics Report"):
    """Export data to Excel format
    Args:
        data: Either a DataFrame or a dictionary of DataFrames
        title: Report title
    """
    try:
        buffer = io.BytesIO()
        wb = Workbook()
        
        # Handle both single DataFrame and dictionary of DataFrames
        if isinstance(data, dict):
            dataframes = data
        else:
            dataframes = {"Data": data}

        # Remove default sheet if we have multiple sheets
        if len(dataframes) > 1:
            wb.remove(wb.active)

        for sheet_name, dataframe in dataframes.items():
            if dataframe.empty:
                continue
                
            # Create worksheet
            if len(dataframes) == 1:
                ws = wb.active
                ws.title = "Statistics"
            else:
                ws = wb.create_sheet(title=sheet_name)

            # Add title
            ws["A1"] = title if len(dataframes) == 1 else f"{title} - {sheet_name}"
            ws["A1"].font = Font(size=16, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center")

            max_col = len(dataframe.columns)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

            # Add generation date
            ws["A2"] = f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws["A2"].font = Font(size=10, italic=True)

            start_row = 4

            # Add headers
            for col_idx, column_name in enumerate(dataframe.columns, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=column_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")

            # Add data rows
            for row_idx, (_, row) in enumerate(dataframe.iterrows(), start_row + 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=str(value))

            # Auto-adjust column widths
            for col_idx in range(1, len(dataframe.columns) + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0

                for row_idx in range(start_row, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        cell_length = len(str(cell_value))
                        if cell_length > max_length:
                            max_length = cell_length

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    except Exception as e:
        st.error(f"Error exporting to Excel: {str(e)}")
        return None

