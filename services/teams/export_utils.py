import io
import streamlit as st
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from database import db_manager
import plotly.express as px


def export_to_csv(dataframe):
    """Export dataframe to CSV format"""
    try:
        return dataframe.to_csv(index=False).encode('utf-8')
    except Exception as e:
        st.error(f"Error exporting to CSV: {str(e)}")
        return None


def export_to_pdf(dataframe, title="Team Statistics Report"):
    """Export dataframe to PDF format"""
    try:
        buffer = io.BytesIO()
        # Utiliser des marges plus petites pour plus d'espace
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)

        # Container for the 'Flowable' objects
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=20,
            alignment=1  # Center alignment
        )

        # Add title
        title_para = Paragraph(title, title_style)
        elements.append(title_para)
        elements.append(Spacer(1, 8))

        # Add generation date
        date_para = Paragraph(f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 12))

        # Prepare data for table avec troncature plus agressive
        data = [list(dataframe.columns)]  # Header
        for _, row in dataframe.iterrows():
            processed_row = []
            for i, cell in enumerate(row):
                cell_str = str(cell)
                # Troncature différente selon la colonne
                if i == 0:  # ID
                    processed_row.append(cell_str[:8])
                elif i == 1:  # Team Name
                    processed_row.append(cell_str[:15])
                elif i == 2:  # Code
                    processed_row.append(cell_str[:10])
                elif i == 3:  # Description
                    processed_row.append(cell_str[:20] + '...' if len(cell_str) > 20 else cell_str)
                elif i == 4:  # Manager
                    processed_row.append(cell_str[:15])
                elif i == 5:  # Member
                    processed_row.append(cell_str[:12])
                elif i == 6:  # Member Email
                    processed_row.append(cell_str[:20] + '...' if len(cell_str) > 20 else cell_str)
                elif i == 7:  # Member Role
                    processed_row.append(cell_str[:10])
                else:  # Dates
                    processed_row.append(cell_str[:12])
            data.append(processed_row)

        # Calculer la largeur disponible
        available_width = A4[0] - 60  # Largeur A4 moins les marges
        col_count = len(dataframe.columns)
        col_width = available_width / col_count

        # Créer le tableau avec largeurs de colonnes spécifiées
        table = Table(data, colWidths=[col_width] * col_count)

        # Style the table
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alignement à gauche pour économiser l'espace
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),  # Police plus petite
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),  # Police encore plus petite pour les données
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    except Exception as e:
        st.error(f"Error exporting to PDF: {str(e)}")
        return None


def export_to_excel(dataframe, title="Team Statistics Report"):
    """Export dataframe to Excel format"""
    try:
        buffer = io.BytesIO()

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Team Statistics"

        # Add title
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Fusionner les cellules pour le titre en fonction du nombre de colonnes
        max_col = len(dataframe.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

        # Add generation date
        ws['A2'] = f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)

        # Add data starting from row 4
        start_row = 4

        # Add headers manually
        for col_idx, column_name in enumerate(dataframe.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Add data rows
        for row_idx, (_, row) in enumerate(dataframe.iterrows(), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value))

        # Auto-adjust column widths - correction de l'erreur
        for col_idx in range(1, len(dataframe.columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0

            # Parcourir toutes les cellules de la colonne pour trouver la largeur max
            for row_idx in range(start_row, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_length = len(str(cell_value))
                    if cell_length > max_length:
                        max_length = cell_length

            # Ajuster la largeur avec une limite max
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    except Exception as e:
        st.error(f"Error exporting to Excel: {str(e)}")
        return None


def create_team_charts(teams_df):
    """Create charts for team visualization"""
    try:
        charts = {}

        if teams_df.empty:
            return charts

        # Chart 1: Teams by Manager
        manager_data = []
        for _, team in teams_df.iterrows():
            team_details = db_manager.get_team_by_id(team['id'])
            manager_name = team_details.get('manager_name', 'No manager') if team_details else 'No manager'
            manager_data.append(manager_name)

        if manager_data:
            manager_counts = pd.Series(manager_data).value_counts()
            charts['managers'] = px.pie(
                values=manager_counts.values,
                names=manager_counts.index,
                title="Teams by Manager"
            )

        # Chart 2: Team sizes distribution
        size_data = []
        for _, team in teams_df.iterrows():
            members = db_manager.get_team_members(team['id'])
            size_data.append(len(members))

        if size_data:
            size_df = pd.DataFrame({'Team Size': size_data})
            charts['sizes'] = px.histogram(
                size_df,
                x='Team Size',
                title="Team Size Distribution",
                nbins=10
            )

        # Chart 3: Teams creation timeline
        if 'created_at' in teams_df.columns:
            timeline_df = teams_df.copy()
            timeline_df['created_at'] = pd.to_datetime(timeline_df['created_at'])
            timeline_df['month'] = timeline_df['created_at'].dt.to_period('M')
            monthly_counts = timeline_df.groupby('month').size().reset_index(name='count')
            monthly_counts['month'] = monthly_counts['month'].astype(str)

            charts['timeline'] = px.line(
                monthly_counts,
                x='month',
                y='count',
                title="Teams Creation Timeline",
                markers=True
            )

        return charts

    except Exception as e:
        st.error(f"Error creating charts: {str(e)}")
        return {}
